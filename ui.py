import ttkbootstrap as ttk
import tkinter as tk
from ttkbootstrap.constants import *
from decimal import Decimal
from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importing sanitized logic and configuration
from calculations import calculate_unit_amount, calculate_subtotal, calculate_total
import config as cfg

# --- Configuration & Safety Fallbacks ---
MOE_ITEMS = getattr(cfg, "MOE_ITEMS", {})
UF_VALUE = getattr(cfg, "UF_VALUE", 420)
DEFAULT_INDEX = getattr(cfg, "DEFAULT_INDEX", 1.6)

def format_currency_ars(value: Decimal) -> str:
    """Formats a Decimal value into a professional ARS currency string ($1.234,56)."""
    integer_part, _, decimal_part = f"{value:.2f}".partition(".")
    formatted_integer = "{:,}".format(int(integer_part)).replace(",", ".")
    return f"${formatted_integer},{decimal_part}"

class CalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Technical Opinions Calculator | Industrial Audit Tool")
        self.root.geometry("1024x768")

        # Header Section
        self._create_header()

        # State Management
        self.uf_value_var = tk.StringVar(value=str(UF_VALUE))
        self.index_var = tk.StringVar(value=str(DEFAULT_INDEX))
        self.quantities = {}
        self.items_flat = {}
        self.item_widgets = []  # Tracking list for UI lifecycle management
        self.subtotal_labels = {}
        
        self.flatten_items()
        self.create_interface()
        self._wire_live_updates()

    def _create_header(self):
        title_frame = ttk.Frame(self.root, bootstyle=PRIMARY)
        title_frame.pack(fill=X, pady=(0, 20))
        ttk.Label(
            title_frame, 
            text="TECHNICAL OPINIONS CALCULATOR", 
            font=("Segoe UI", 18, "bold"),
            bootstyle=INVERSE
        ).pack(pady=15)

    def _wire_live_updates(self):
        """Observer pattern: triggers recalculations on config changes."""
        self.uf_value_var.trace_add("write", lambda *_: self.calculate_final_total())
        self.index_var.trace_add("write", lambda *_: self.calculate_final_total())

    def _decimal_from_var(self, var: tk.StringVar, default="0"):
        """Sanitizes UI string input into Decimal for financial precision."""
        try:
            return Decimal(var.get().strip().replace(",", "."))
        except Exception:
            return Decimal(default)

    def flatten_items(self):
        """Maps hierarchical categories into a flat reference dictionary."""
        for category, items in MOE_ITEMS.items():
            if isinstance(items, dict):
                for item_name, uf_val in items.items():
                    self.items_flat[item_name] = uf_val
            else:
                self.items_flat[category] = items

    def create_interface(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=BOTH, expand=True, padx=20, pady=5)

        main_frame.columnconfigure(0, minsize=300, weight=0)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)

        # --- Sidebar: Category Selection ---
        cat_frame = ttk.LabelFrame(main_frame, text=" 1. Select Categories ", padding=10)
        cat_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
        
        self.category_listbox = tk.Listbox(
            cat_frame, selectmode=tk.MULTIPLE, exportselection=False,
            font=("Segoe UI", 10), bd=0, highlightthickness=0
        )
        self.category_listbox.pack(side=LEFT, fill=BOTH, expand=True)

        scrollbar = ttk.Scrollbar(cat_frame, orient=VERTICAL, command=self.category_listbox.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.category_listbox.config(yscrollcommand=scrollbar.set)

        for cat in sorted(MOE_ITEMS.keys()):
            self.category_listbox.insert(tk.END, cat)
        self.category_listbox.bind("<<ListboxSelect>>", self.update_item_rows)

        # --- Main Workspace: Dynamic Item Grid ---
        work_frame = ttk.LabelFrame(main_frame, text=" 2. Set Quantities & Review Audit ", padding=10)
        work_frame.grid(row=1, column=1, sticky="nsew")

        # Scrollable container for items
        self.canvas = tk.Canvas(work_frame, highlightthickness=0)
        self.items_inner_frame = ttk.Frame(self.canvas)
        self.scroll_y = ttk.Scrollbar(work_frame, orient=VERTICAL, command=self.canvas.yview)
        
        self.canvas.configure(yscrollcommand=self.scroll_y.set)
        self.scroll_y.pack(side=RIGHT, fill=Y)
        self.canvas.pack(side=LEFT, fill=BOTH, expand=True)
        self.canvas_window = self.canvas.create_window((0,0), window=self.items_inner_frame, anchor="nw")

        self.items_inner_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        # --- Footer: Configuration & Results ---
        footer = ttk.Frame(self.root, padding=20)
        footer.pack(fill=X)

        # Config Inputs
        config_box = ttk.Frame(footer)
        config_box.pack(side=LEFT)
        ttk.Label(config_box, text="UF Value:").pack(side=LEFT, padx=5)
        ttk.Entry(config_box, textvariable=self.uf_value_var, width=10).pack(side=LEFT, padx=5)
        ttk.Label(config_box, text="Index:").pack(side=LEFT, padx=5)
        ttk.Entry(config_box, textvariable=self.index_var, width=10).pack(side=LEFT, padx=5)

        # Results & Actions
        actions_box = ttk.Frame(footer)
        actions_box.pack(side=RIGHT)
        self.total_label = ttk.Label(actions_box, text="Total: $0,00", font=("Segoe UI", 16, "bold"), bootstyle=SUCCESS)
        self.total_label.pack(side=TOP, anchor="e", pady=(0, 10))
        
        ttk.Button(actions_box, text="Export Excel (.xlsx)", command=self.export_excel, bootstyle=SUCCESS).pack(side=RIGHT, padx=5)
        ttk.Button(actions_box, text="Reset", command=self.reset, bootstyle=DANGER).pack(side=RIGHT, padx=5)

    def update_item_rows(self, *args):
        """Lifecycle: Clears and rebuilds the grid ensuring object destruction."""
        for widget in self.item_widgets:
            if hasattr(widget, 'destroy'):
                widget.destroy()
        
        self.item_widgets.clear()
        self.quantities.clear()
        self.subtotal_labels.clear()

        selected_indices = self.category_listbox.curselection()
        
        # Grid Headers
        headers = ["Item Description", "UF", "Quantity", "Subtotal (ARS)"]
        for col, text in enumerate(headers):
            h_lbl = ttk.Label(self.items_inner_frame, text=text, font=("Segoe UI", 9, "bold"))
            h_lbl.grid(row=0, column=col, padx=15, pady=10, sticky="w")
            self.item_widgets.append(h_lbl)

        row = 1
        for idx in selected_indices:
            category = self.category_listbox.get(idx)
            items = MOE_ITEMS.get(category, {})
            
            if isinstance(items, dict):
                for item_name, uf_val in sorted(items.items()):
                    self._create_row(row, item_name, uf_val)
                    row += 1
            else:
                self._create_row(row, category, items)
                row += 1

        self.calculate_final_total()

    def _create_row(self, row, name, uf):
        """Internal helper to build dynamic UI rows and track widgets."""
        lbl_name = ttk.Label(self.items_inner_frame, text=name, wraplength=400)
        lbl_name.grid(row=row, column=0, sticky="w", padx=15, pady=5)
        
        lbl_uf = ttk.Label(self.items_inner_frame, text=f"{uf} UF")
        lbl_uf.grid(row=row, column=1, padx=15)
        
        q_var = tk.IntVar(value=0)
        sb_qty = ttk.Spinbox(self.items_inner_frame, from_=0, to=999, width=8, textvariable=q_var)
        sb_qty.grid(row=row, column=2, padx=15)
        q_var.trace_add("write", lambda *_: self.calculate_final_total())
        
        lbl_sub = ttk.Label(self.items_inner_frame, text="$0,00", font=("Consolas", 10))
        lbl_sub.grid(row=row, column=3, padx=15, sticky="e")

        # CRITICAL FIX: Track all widgets for proper UI cleanup
        self.item_widgets.extend([lbl_name, lbl_uf, sb_qty, lbl_sub])
        self.quantities[name] = q_var
        self.subtotal_labels[name] = lbl_sub

    def calculate_final_total(self):
        """Financial engine: aggregates subtotals with high precision."""
        try:
            subtotals_list = []
            current_uf_price = self._decimal_from_var(self.uf_value_var)
            
            for name, q_var in self.quantities.items():
                qty = q_var.get()
                uf_units = Decimal(str(self.items_flat[name]))
                
                # Using the translated calculation module
                unit_amount = calculate_unit_amount(uf_units, current_uf_price)
                subtotal = calculate_subtotal(unit_amount, qty)
                
                subtotals_list.append(subtotal)
                self.subtotal_labels[name].config(text=format_currency_ars(subtotal))

            current_index = self._decimal_from_var(self.index_var, default="1")
            final_res = calculate_total(subtotals_list, current_index) if subtotals_list else Decimal("0")
            
            self.total_label.config(text=f"Total: {format_currency_ars(final_res)}")
        except Exception:
            pass 

    def reset(self):
        """Clears all user input to default state."""
        for v in self.quantities.values():
            v.set(0)
        self.calculate_final_total()

    def export_excel(self):
        """Generates a Professional .xlsx Report with formatting."""
        file_path = filedialog.asksaveasfilename(
            title="Export Professional Audit",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            # 1. Setup Workbook and Styles
            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Report"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid") # Dark Blue
            total_font = Font(bold=True, color="27ae60", size=12) # Green for money
            center_align = Alignment(horizontal="center")
            
            # 2. Write Headers
            headers = ["Item Description", "UF Units", "Quantity", "Subtotal (ARS)"]
            ws.append(headers)
            
            # Apply Header Styles
            for col_num, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # 3. Write Data
            current_uf = self.uf_value_var.get()
            
            for name, q_var in self.quantities.items():
                qty = q_var.get()
                if qty > 0:
                    # Get clean values for Excel math
                    sub_text = self.subtotal_labels[name].cget("text")
                    # Clean currency formatting for Excel number format (optional logic depending on your needs)
                    
                    ws.append([name, self.items_flat[name], qty, sub_text])

            # 4. Add Summary Section at the bottom
            ws.append([]) # Empty row
            ws.append(["CONFIGURATION & TOTALS"])
            ws["A" + str(ws.max_row)].font = Font(bold=True)
            
            ws.append(["Configured UF Value:", current_uf])
            ws.append(["Applied Index:", self.index_var.get()])
            
            # Final Total Row
            total_text = self.total_label.cget("text") # e.g. "Total: $1.200,00"
            ws.append(["FINAL AUDIT TOTAL", "", "", total_text])
            
            # Style the Total
            last_row = ws.max_row
            ws[f"D{last_row}"].font = total_font
            ws[f"D{last_row}"].alignment = Alignment(horizontal="right")

            # 5. Auto-adjust Column Widths (The "Magic" touch)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Save
            wb.save(file_path)
            messagebox.showinfo("Success", f"Professional Report saved to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Error generating Excel:\n{str(e)}")